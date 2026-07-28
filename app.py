import io
import re
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime
import streamlit as st
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIGURAÇÃO DA PÁGINA
# ==========================================
st.set_page_config(page_title="Extrator TIC Trens", page_icon="🚆", layout="centered")
st.title("Cruzamento de Pedidos e NFs - TIC Trens")
st.markdown("Faça o upload dos arquivos extraídos do SAP para gerar a planilha mestra.")

# ==========================================
# MÓDULO 1: LEITURA DE NOTAS FISCAIS (XML)
# ==========================================
def extrair_dados_xml(arquivo_xml):
    dados_extraidos = []
    try:
        tree = ET.parse(arquivo_xml)
        root = tree.getroot()
        ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}
        
        infNFe = root.find('.//ns:infNFe', ns)
        if infNFe is None:
            return []

        ide = infNFe.find('ns:ide', ns)
        numero_nf = "Não consta"
        data_emissao = None
        
        if ide is not None:
            tag_nNF = ide.find('ns:nNF', ns)
            if tag_nNF is not None:
                numero_nf = tag_nNF.text
                
            tag_dhEmi = ide.find('ns:dhEmi', ns)
            if tag_dhEmi is not None:
                data_bruta = tag_dhEmi.text.split('T')[0]
                try:
                    data_emissao = datetime.strptime(data_bruta, '%Y-%m-%d').date()
                except ValueError:
                    data_emissao = None

        numero_pv = "Não consta"
        numero_pc = "Não consta"
        
        infAdic = infNFe.find('ns:infAdic', ns)
        if infAdic is not None:
            tag_infCpl = infAdic.find('ns:infCpl', ns)
            if tag_infCpl is not None and tag_infCpl.text:
                texto_obs = tag_infCpl.text
                match_pc = re.search(r'(?:PC|Pedido)\s*:?\s*(\d{10})\b', texto_obs, re.IGNORECASE)
                if match_pc:
                    numero_pc = match_pc.group(1)
                match_pv = re.search(r'\bPV\s*(\d{4,5})\b', texto_obs)
                if match_pv:
                    numero_pv = match_pv.group(1)

        for det in infNFe.findall('ns:det', ns):
            prod = det.find('ns:prod', ns)
            if prod is not None:
                tag_cProd = prod.find('ns:cProd', ns)
                codigo = tag_cProd.text if tag_cProd is not None else ""
                
                tag_xProd = prod.find('ns:xProd', ns)
                produto = tag_xProd.text if tag_xProd is not None else ""
                
                match_cc = re.search(r'\(\s*(?:CC\s*)?(\d{6})\s*\)', produto, re.IGNORECASE)
                codigo_cliente = match_cc.group(1) if match_cc else ""
                
                tag_qCom = prod.find('ns:qCom', ns)
                quantidade = float(tag_qCom.text) if tag_qCom is not None else 0.0
                
                tag_vUnCom = prod.find('ns:vUnCom', ns)
                preco_un = float(tag_vUnCom.text) if tag_vUnCom is not None else 0.0
                
                tag_vProd = prod.find('ns:vProd', ns)
                preco_tot = float(tag_vProd.text) if tag_vProd is not None else 0.0
                
                pc_do_item = numero_pc
                if pc_do_item == "Não consta":
                    tag_xPed = prod.find('ns:xPed', ns)
                    if tag_xPed is not None and tag_xPed.text:
                        match_xped = re.search(r'(?:PC)?\s*(\d{10})\b', tag_xPed.text, re.IGNORECASE)
                        if match_xped:
                            pc_do_item = match_xped.group(1)
                
                dados_extraidos.append({
                    "Código": codigo,
                    "Produto": produto,
                    "Código do Cliente": codigo_cliente,
                    "Quantidade": quantidade,
                    "Preço unitário": preco_un,
                    "Preço total": preco_tot,
                    "Número da NF": numero_nf,
                    "Data de emissão": data_emissao,
                    "Número do PV": numero_pv,
                    "Número do PC": pc_do_item
                })
    except ET.ParseError:
        pass
    return dados_extraidos


# ==========================================
# MÓDULO 2: LEITURA DE PEDIDOS (TXT E XLSX)
# ==========================================
def extrair_dados_txt(arquivo_txt):
    dados_extraidos = []
    texto_completo = arquivo_txt.getvalue().decode('utf-8', errors='ignore')
    linhas = texto_completo.splitlines(keepends=True)
    
    data_global = None
    match_gd = re.search(r'Data da remessa\s+(\d{2}\.\d{2}\.\d{4})', texto_completo, re.IGNORECASE)
    if match_gd:
        data_global = match_gd.group(1)

    pedido_compra = "Não encontrado"
    match_pedido = re.search(r'Número do Documento\s+(\d{10})', texto_completo, re.IGNORECASE)
    if match_pedido:
        pedido_compra = match_pedido.group(1)
    else:
        match_fallback = re.search(r'\b(4[56]\d{8})\b', texto_completo)
        if match_fallback:
            pedido_compra = match_fallback.group(1)

    i = 0
    while i < len(linhas):
        linha = linhas[i].strip()
        padrao_item = r'(?:\b\d+\s+)?(\d{6})\s+(\d+(?:\.\d+)?,\d{2})\s+([A-Za-z]{1,4})\s+([\d.,]+)(?:\s*BRL\s*/\s*[A-Za-z]+)?\s+([\d.,]+)\s*BRL'
        match = re.search(padrao_item, linha)
        
        if match:
            codigo_cliente = match.group(1)
            qtd_str = match.group(2)
            um = match.group(3)
            preco_unit_str = match.group(4)
            preco_tot_str = match.group(5)
            
            descricao_partes = []
            j = i + 1
            while j < len(linhas):
                proxima_linha = linhas[j].strip()
                linha_lower = proxima_linha.lower()
                
                if re.search(padrao_item, proxima_linha) or \
                   "data da remessa" in linha_lower or \
                   "texto longo:" in linha_lower or \
                   "marca:" in linha_lower or \
                   proxima_linha == "0,00" or \
                   "offset" in linha_lower or \
                   "valor total" in linha_lower or \
                   "_______________" in proxima_linha or \
                   proxima_linha.startswith("***"):
                    break
                
                if not proxima_linha or \
                   "página" in linha_lower or \
                   "page " in linha_lower or \
                   "preço líquido" in linha_lower or \
                   "preco liquido" in linha_lower or \
                   "item material/descrição" in linha_lower or \
                   re.match(r'^4[56]\d{8}$', proxima_linha) or \
                   re.match(r'^BRL\s*/\s*[A-Za-z]+$', proxima_linha) or \
                   (re.match(r'^\d+$', proxima_linha) and len(proxima_linha) < 5): 
                    j += 1
                    continue
                
                descricao_partes.append(proxima_linha)
                j += 1
                
            produto = " ".join(descricao_partes).strip()
            if not produto:
                produto = "Produto sem descrição"
                
            data_item_str = data_global
            for k in range(i, min(i + 15, len(linhas))):
                match_data = re.search(r'Data da remessa:\s*(?:Dia\s+)?(\d{2}\.\d{2}\.\d{4})', linhas[k], re.IGNORECASE)
                if match_data:
                    data_item_str = match_data.group(1)
                    break

            try: quantidade = float(qtd_str.replace('.', '').replace(',', '.'))
            except ValueError: quantidade = 0.0
                
            try: preco_unit = float(preco_unit_str.replace('.', '').replace(',', '.'))
            except ValueError: preco_unit = 0.0
                
            try: preco_total = float(preco_tot_str.replace('.', '').replace(',', '.'))
            except ValueError: preco_total = 0.0
            
            data_obj = None
            if data_item_str:
                try: data_obj = datetime.strptime(data_item_str, '%d.%m.%Y').date()
                except ValueError: pass
            
            dados_extraidos.append({
                "Pedido de Compras": pedido_compra,
                "Código do Cliente": codigo_cliente,
                "Produto": produto,
                "Quantidade": quantidade,
                "Unidade de Medida": um,
                "Preço unitário": preco_unit,
                "Preço total": preco_total,
                "Data da remessa": data_obj
            })
            i = j - 1
        i += 1
    return dados_extraidos


def extrair_dados_xlsx(arquivo_xlsx):
    df = pd.read_excel(arquivo_xlsx, header=None)
    
    header_idx = None
    for i, row in df.iterrows():
        row_strs = [str(val).strip().upper() for val in row.values if pd.notna(val)]
        if any(kw in row_strs for kw in ['CÓDIGO MATERIAL', 'MATERIAL', 'Nº MATERIAL', 'CÓDIGO', 'MATERIAL/DESCRIÇÃO', 'CÓD. CLIENTE']):
            header_idx = i
            break
            
    if header_idx is None:
        return []
        
    df.columns = [str(val).strip().upper() if pd.notna(val) else f"UNNAMED_{idx}" for idx, val in enumerate(df.iloc[header_idx])]
    df = df.iloc[header_idx+1:].copy()
    
    nome_arquivo = arquivo_xlsx.name
    match_pedido = re.search(r'\b(4[56]\d{8})\b', nome_arquivo)
    pedido_compra_arquivo = match_pedido.group(1) if match_pedido else "Não encontrado"
    
    col_codigo = next((c for c in df.columns if c in ['CÓDIGO MATERIAL', 'MATERIAL', 'Nº MATERIAL', 'CÓDIGO', 'MATERIAL/DESCRIÇÃO', 'CÓD. CLIENTE']), None)
    col_desc = next((c for c in df.columns if c in ['DESCRIÇÃO', 'TEXTO BREVE', 'TEXTO BREVE DO MATERIAL', 'PRODUTO']), None)
    col_texto = next((c for c in df.columns if c in ['TEXTO LONGO']), None)
    col_qtd = next((c for c in df.columns if c in ['QTD', 'QUANTIDADE', 'QTD. PEDIDO', 'QTD.']), None)
    col_um = next((c for c in df.columns if c in ['UND', 'UM', 'UNIDADE', 'UNID. MEDIDA']), None)
    col_pr_unit = next((c for c in df.columns if c in ['PR.UNIT.', 'PREÇO LÍQUIDO', 'VALOR UNITÁRIO', 'VLR. UNIT.', 'UNIDADE DE PREÇO', 'PREÇO UNITÁRIO (R$)', 'PREÇO UNITÁRIO']), None)
    col_pr_tot = next((c for c in df.columns if c in ['PR. TOTAL', 'VALOR LÍQUIDO', 'VLR. TOTAL', 'VALOR TOTAL', 'PREÇO TOTAL (R$)', 'PREÇO TOTAL']), None)
    col_data = next((c for c in df.columns if c in ['PRAZO DE ENTREGA', 'DATA DE REMESSA', 'DATA DE ENTREGA', 'DATA DA REMESSA', 'DATA REQUISIÇÃO']), None)
    col_pedido = next((c for c in df.columns if c in ['Nº PEDIDO', 'PEDIDO', 'PEDIDO DE COMPRAS']), None)
    
    dados_extraidos = []
    for _, row in df.iterrows():
        if not col_codigo or pd.isna(row.get(col_codigo)):
            continue
            
        codigo_cliente = str(row[col_codigo]).replace('.0', '').strip()
        if not re.match(r'^\d+$', codigo_cliente):
            continue
            
        desc = str(row.get(col_desc, '')) if col_desc else ''
        desc = desc.strip()
        if desc.lower() == 'nan': desc = ''
        
        if not desc:
            texto_longo = str(row.get(col_texto, '')) if col_texto else ''
            texto_longo = texto_longo.strip()
            if texto_longo.lower() == 'nan': texto_longo = ''
            produto = texto_longo or "Produto sem descrição"
        else:
            produto = desc
            
        quantidade = pd.to_numeric(row.get(col_qtd), errors='coerce') if col_qtd else 0.0
        
        um = str(row.get(col_um, '')) if col_um else ''
        um = um.strip()
        if um.lower() == 'nan': um = ''
            
        preco_unit = pd.to_numeric(row.get(col_pr_unit), errors='coerce') if col_pr_unit else 0.0
        preco_total = pd.to_numeric(row.get(col_pr_tot), errors='coerce') if col_pr_tot else 0.0
        
        data_remessa_raw = row.get(col_data) if col_data else None
        data_obj = None
        if pd.notna(data_remessa_raw):
            if isinstance(data_remessa_raw, datetime) or "Timestamp" in str(type(data_remessa_raw)):
                data_obj = data_remessa_raw.date()
            else:
                match_data = re.search(r'(\d{2})[./-](\d{2})[./-](\d{4})', str(data_remessa_raw))
                if match_data:
                    try: data_obj = datetime.strptime(f"{match_data.group(1)}/{match_data.group(2)}/{match_data.group(3)}", '%d/%m/%Y').date()
                    except ValueError: pass
        
        pedido_linha = pedido_compra_arquivo
        if col_pedido and pd.notna(row.get(col_pedido)):
            val_ped = str(row[col_pedido]).replace('.0', '').strip()
            if re.match(r'^4[56]\d{8}$', val_ped):
                pedido_linha = val_ped
        
        dados_extraidos.append({
            "Pedido de Compras": pedido_linha,
            "Código do Cliente": codigo_cliente,
            "Produto": produto,
            "Quantidade": float(quantidade) if pd.notna(quantidade) else 0.0,
            "Unidade de Medida": um,
            "Preço unitário": float(preco_unit) if pd.notna(preco_unit) else 0.0,
            "Preço total": float(preco_total) if pd.notna(preco_total) else 0.0,
            "Data da remessa": data_obj
        })
    return dados_extraidos


# ==========================================
# INTERFACE DO USUÁRIO E CRUZAMENTO
# ==========================================
st.subheader("1. Envie os Arquivos")
col1, col2 = st.columns(2)

with col1:
    arquivos_nfs = st.file_uploader("Notas Fiscais (XML)", type=['xml'], accept_multiple_files=True)
with col2:
    arquivos_pcs = st.file_uploader("Pedidos de Compra (TXT, XLSX)", type=['txt', 'xlsx'], accept_multiple_files=True)

if st.button("Cruzar Dados e Gerar Planilha", type="primary", use_container_width=True):
    if not arquivos_nfs and not arquivos_pcs:
        st.warning("Por favor, envie pelo menos um arquivo para processar.")
    else:
        with st.spinner("Processando dados e gerando planilha..."):
            todos_nfs = []
            todos_pcs = []
            
            for nf in arquivos_nfs:
                todos_nfs.extend(extrair_dados_xml(nf))
                
            for pc in arquivos_pcs:
                if pc.name.lower().endswith('.txt'):
                    todos_pcs.extend(extrair_dados_txt(pc))
                elif pc.name.lower().endswith('.xlsx'):
                    todos_pcs.extend(extrair_dados_xlsx(pc))

            df_nfs = pd.DataFrame(todos_nfs) if todos_nfs else pd.DataFrame()
            df_pcs = pd.DataFrame(todos_pcs) if todos_pcs else pd.DataFrame()
            
            if df_pcs.empty and df_nfs.empty:
                st.error("Nenhum dado pôde ser extraído dos arquivos informados.")
            else:
                df_cruzamento = pd.DataFrame()
                
                if not df_pcs.empty:
                    df_pcs['Código do Cliente'] = df_pcs['Código do Cliente'].astype(str).str.strip()
                    df_pcs['Quantidade'] = pd.to_numeric(df_pcs['Quantidade'], errors='coerce').fillna(0)
                
                if not df_nfs.empty:
                    df_nfs['Código do Cliente'] = df_nfs['Código do Cliente'].astype(str).str.strip()
                    df_nfs['Quantidade'] = pd.to_numeric(df_nfs['Quantidade'], errors='coerce').fillna(0)

                if not df_pcs.empty:
                    codigos_unicos = df_pcs['Código do Cliente'].unique()
                    dados_cruzamento = []
                    
                    for codigo in codigos_unicos:
                        if codigo == 'nan' or codigo == '':
                            continue
                            
                        pc_data = df_pcs[df_pcs['Código do Cliente'] == codigo]
                        descricao = pc_data['Produto'].iloc[0] if not pc_data.empty else ""
                        um = pc_data['Unidade de Medida'].iloc[0] if not pc_data.empty else ""
                        qtd_solicitada = pc_data['Quantidade'].sum()
                        
                        pcs_lista_pc = pc_data['Pedido de Compras'].dropna().astype(str).str.strip().unique()
                        pc_str = ", ".join([p for p in pcs_lista_pc if p.lower() not in ['nan', 'não encontrado', '']])
                        
                        qtd_faturada = 0
                        nfs_str, pvs_str = "", ""
                        
                        if not df_nfs.empty:
                            nf_data = df_nfs[df_nfs['Código do Cliente'] == codigo]
                            qtd_faturada = nf_data['Quantidade'].sum()
                            
                            nfs_lista = nf_data['Número da NF'].dropna().astype(str).str.strip().unique()
                            nfs_str = ", ".join([nf for nf in nfs_lista if nf.lower() not in ['nan', 'não consta', '']])
                            
                            pvs_lista = nf_data['Número do PV'].dropna().astype(str).str.strip().unique()
                            pvs_str = ", ".join([pv for pv in pvs_lista if pv.lower() not in ['nan', 'não consta', '']])
                            
                        saldo = qtd_solicitada - qtd_faturada
                        
                        dados_cruzamento.append({
                            "Código do Cliente": codigo,
                            "Descrição do Produto": descricao,
                            "Unidade de Medida": um,
                            "Quantidade Faturada": qtd_faturada,
                            "Quantidade Solicitada": qtd_solicitada,
                            "Saldo": saldo,
                            "Notas Fiscais": nfs_str,
                            "Pedido de Venda": pvs_str,
                            "Pedido de Compra": pc_str
                        })
                        
                    df_cruzamento = pd.DataFrame(dados_cruzamento)

                if not df_nfs.empty:
                    colunas_nfs = ["Código", "Produto", "Código do Cliente", "Quantidade", "Preço unitário", "Preço total", "Número da NF", "Data de emissão", "Número do PV", "Número do PC"]
                    df_nfs = df_nfs[[c for c in colunas_nfs if c in df_nfs.columns]]
                    
                if not df_pcs.empty:
                    colunas_pcs = ["Pedido de Compras", "Código do Cliente", "Produto", "Quantidade", "Unidade de Medida", "Preço unitário", "Preço total", "Data da remessa"]
                    df_pcs = df_pcs[[c for c in colunas_pcs if c in df_pcs.columns]]

                # Cria o arquivo em memória para download
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl', datetime_format='DD/MM/YYYY') as writer:
                    if not df_nfs.empty:
                        df_nfs.to_excel(writer, index=False, sheet_name='Notas Fiscais')
                        worksheet_nf = writer.sheets['Notas Fiscais']
                        for row in range(2, len(df_nfs) + 2):
                            worksheet_nf.cell(row=row, column=5).number_format = '"R$" #,##0.00'
                            worksheet_nf.cell(row=row, column=6).number_format = '"R$" #,##0.00'

                    if not df_pcs.empty:
                        df_pcs.to_excel(writer, index=False, sheet_name='Pedido de Compras')
                        worksheet_pc = writer.sheets['Pedido de Compras']
                        for row in range(2, len(df_pcs) + 2):
                            worksheet_pc.cell(row=row, column=6).number_format = '"R$" #,##0.00'
                            worksheet_pc.cell(row=row, column=7).number_format = '"R$" #,##0.00'
                            
                    if not df_cruzamento.empty:
                        df_cruzamento.to_excel(writer, index=False, sheet_name='Cruzamento')
                        
                    workbook = writer.book
                    
                    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    fill_amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                    for sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        ws.freeze_panes = "A2"
                        
                        for col in ws.columns:
                            max_len = 0
                            col_letter = get_column_letter(col[0].column)
                            for cell in col:
                                if cell.value is not None:
                                    val_str = str(cell.value)
                                    if len(val_str) > max_len:
                                        max_len = len(val_str)
                            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                            
                        if sheet_name == 'Cruzamento':
                            saldo_col_idx = None
                            for col_idx, header_cell in enumerate(ws[1], 1):
                                if header_cell.value == 'Saldo':
                                    saldo_col_idx = col_idx
                                    break
                            
                            if saldo_col_idx:
                                for row in range(2, ws.max_row + 1):
                                    saldo_val = ws.cell(row=row, column=saldo_col_idx).value
                                    try:
                                        saldo_num = float(saldo_val) if saldo_val is not None else 0.0
                                    except ValueError:
                                        saldo_num = 0.0
                                        
                                    linha_fill = None
                                    if saldo_num > 0:
                                        linha_fill = fill_vermelho
                                    elif saldo_num < 0:
                                        linha_fill = fill_amarelo
                                    else:
                                        linha_fill = fill_verde
                                        
                                    if linha_fill:
                                        for c in range(1, ws.max_column + 1):
                                            ws.cell(row=row, column=c).fill = linha_fill

                st.success("Operação concluída com sucesso!")
                
                # Exibe o botão de download com o arquivo gerado
                st.download_button(
                    label="Baixar Planilha de Cruzamento",
                    data=output.getvalue(),
                    file_name="Planilha_Mestra_Cruzamento.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
